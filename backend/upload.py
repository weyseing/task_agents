"""Parse uploaded csv/xlsx into the JSON shape stored in R2.

Single-sheet inputs → flat {columns, rows}.
Multi-sheet xlsx   → {sheets: [{name, columns, rows}, ...]}.
"""

import csv
import io


# 20MB — large enough for real workbooks, small enough that one user can't
# blow through a Cloud Run instance's memory.
MAX_UPLOAD_BYTES = 20 * 1024 * 1024

ALLOWED_EXTS = {"csv", "xlsx"}


def detect_type(filename: str) -> str | None:
    """Return 'csv' or 'xlsx', or None if extension is not allowed."""
    if not filename:
        return None
    lower = filename.rsplit("/", 1)[-1].lower()
    if lower.endswith(".csv"):
        return "csv"
    if lower.endswith(".xlsx"):
        return "xlsx"
    return None


def _to_str(v) -> str:
    if v is None:
        return ""
    return str(v)


def parse_csv(blob: bytes) -> dict:
    """Parse a csv into {columns, rows}. First non-empty row is the header."""
    try:
        text = blob.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = blob.decode("latin-1", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    while rows and not any(c.strip() for c in rows[0]):
        rows.pop(0)
    if not rows:
        return {"columns": [], "rows": []}
    columns = [_to_str(c).strip() for c in rows[0]]
    width = len(columns)
    data: list[list[str]] = []
    for r in rows[1:]:
        row = [_to_str(c) for c in r]
        if len(row) < width:
            row += [""] * (width - len(row))
        elif len(row) > width:
            row = row[:width]
        data.append(row)
    return {"columns": columns, "rows": data}


def _strip_trailing_empty_columns(columns: list[str], rows: list[list[str]]) -> tuple[list[str], list[list[str]]]:
    """openpyxl reports the full max_column for every sheet, which often
    includes trailing empties when a sheet was edited then trimmed. Drop
    columns that have a blank header AND no data."""
    if not columns:
        return columns, rows
    keep = len(columns)
    while keep > 0:
        i = keep - 1
        header_blank = not columns[i].strip()
        col_blank = all((i >= len(r) or not r[i].strip()) for r in rows)
        if header_blank and col_blank:
            keep -= 1
        else:
            break
    if keep == len(columns):
        return columns, rows
    return columns[:keep], [r[:keep] for r in rows]


def _strip_trailing_empty_rows(rows: list[list[str]]) -> list[list[str]]:
    while rows and all(not (c or "").strip() for c in rows[-1]):
        rows.pop()
    return rows


def parse_xlsx(blob: bytes) -> dict:
    """Parse an xlsx workbook. Returns:
       - flat {columns, rows} for single-sheet workbooks (so existing
         agent tools that don't pass `sheet` keep working unchanged)
       - {sheets: [...]} for multi-sheet workbooks
    """
    # Imported lazily so the rest of the backend doesn't pay the cost when
    # uploads are unused.
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    sheets_out: list[dict] = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            sheets_out.append({"name": name, "columns": [], "rows": []})
            continue
        columns = [_to_str(c).strip() for c in (header_row or [])]
        data: list[list[str]] = []
        width = len(columns)
        for r in rows_iter:
            row = [_to_str(c) for c in (r or [])]
            if len(row) < width:
                row += [""] * (width - len(row))
            elif len(row) > width:
                row = row[:width]
            data.append(row)
        data = _strip_trailing_empty_rows(data)
        columns, data = _strip_trailing_empty_columns(columns, data)
        sheets_out.append({"name": name, "columns": columns, "rows": data})

    if not sheets_out:
        return {"columns": [], "rows": []}
    if len(sheets_out) == 1:
        s = sheets_out[0]
        return {"columns": s["columns"], "rows": s["rows"]}
    return {"sheets": sheets_out}


def parse_upload(filename: str, blob: bytes) -> tuple[str, dict]:
    """Detect type and parse. Returns (file_type, content_dict).
    Raises ValueError on unsupported types or oversized blobs."""
    if len(blob) > MAX_UPLOAD_BYTES:
        raise ValueError(
            f"file too large ({len(blob)} bytes); limit is {MAX_UPLOAD_BYTES}"
        )
    ftype = detect_type(filename)
    if ftype is None:
        raise ValueError("only .csv and .xlsx uploads are supported")
    if ftype == "csv":
        return "csv", parse_csv(blob)
    return "xlsx", parse_xlsx(blob)
